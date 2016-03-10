#!/usr/bin/python
# -*- coding: utf-8 -*-
__author__ = 'Yuri.Perin'


import cv2
from openpyxl import load_workbook
from time import sleep
import threading
import os
import sys
from json import load
from shutil import copy


class KThread(threading.Thread):
    """A subclass of threading.Thread, with a kill()
    method."""
    def __init__(self, *args, **keywords):
        threading.Thread.__init__(self, *args, **keywords)
        self.killed = False

    def start(self):
        """Start the thread."""
        self.__run_backup = self.run
        self.run = self.__run  # Force the Thread to install our trace.
        threading.Thread.start(self)

    def __run(self):
        """Hacked run function, which installs the
        trace."""
        sys.settrace(self.globaltrace)
        self.__run_backup()
        self.run = self.__run_backup

    def globaltrace(self, frame, why, arg):
        if why == 'call':
            return self.localtrace
        else:
            return None

    def localtrace(self, frame, why, arg):
        if self.killed:
            if why == 'line':
                raise SystemExit()
            return self.localtrace

    def kill(self):
        self.killed = True


class Grabber:
    """GRABBER"""
    def __init__(self, app):
        self.app = app
        self.cap = cv2.VideoCapture(0)
        sleep(1)
        self.cap.set(3, 720)
        self.cap.set(4, 576)
        self.status = "open"

    def show(self):
        while 1:
            s, img = self.cap.read()
            if s:
                cv2.imshow("Capturing", img)
                if cv2.waitKey(1) == 27:
                    break


class Display:
    def __init__(self, app):
        self.app = app

    def display(self, ret, frame):
        if ret:
            cv2.imshow("Capturing", frame)
            if cv2.waitKey(1) == 27:
                cv2.destroyAllWindows()


class FrameStorage:

    def __init__(self, app):
        self.app = app
        self.frame = None
        with open(os.path.join(self.app.domain, "Configuration", "config_syst.json")) as info:
            self._chip = load(info)["CHIP"][self.app.stb_model]
        self._reference_dir = os.path.join(self.app.domain, "RefPictures", self._chip, self.app.test, self.app.testscript)
        self.compare_result = 0

    def set(self, ret, frame):
        if ret:
            self._frame = frame

    def get(self):
        return self._frame

    def rgb2gray_cv(self, image):
        rgb = cv2.imread(image)
        r, g, b = rgb[:, :, 0], rgb[:, :, 1], rgb[:, :, 2]
        gray = 0.21 * r + 0.72 * g + 0.07 * b
        return gray

    def report_to_xlsx(self, testcase, col, report, result):
        rep = load_workbook(report)
        sheet = rep.get_sheet_by_name("BBT_results")
        c = sheet.cell(row=testcase, column=col)
        c.value = result
        rep.save(report)

    def check_result(self, testcase):
        snapshot = self._frame
        ref_picture = os.path.join(self._reference_dir, "%s_%s.bmp" % (self.app.testscript, testcase))
        test_picture = os.path.join(self.app.domain, "Temp", "DUT_snapshot.bmp")
        cv2.imwrite(test_picture, snapshot)
        self.app.write_log("[GRAB]  Take DUT_snapshot.bmp for testcase %s" % testcase)  # logger - take the snapshot
        print("[GRAB]  Take DUT_snapshot.bmp for testcase %s" % testcase)  # print - take the snapshot
        try:
            dif = abs(self.rgb2gray_cv(ref_picture) - self.rgb2gray_cv(test_picture))
            result = int((255.0 - dif.max())/255*100)
        except:
            result = 0
            dif = None
        self.compare_result = result
        self.app.write_log("[GRAB]  Result = %s" % result)  # logger - fill the xlsx report
        print("[GRAB]  Result = %s" % result)  # print - fill the xlsx report
        if result > 90:
            conclusion = "SUCCESS"
        else:
            conclusion = "FAILED"
            try:
                copy(test_picture, os.path.join(self.app.report_dir, "%s_%s.bmp" % (self.app.testscript, testcase)))
                cv2.imwrite(os.path.join(self.app.report_dir, "%s_%s_dif.jpg" % (self.app.testscript, testcase)), dif)
                return result
            except:
                return result
        self.app.write_log("[GRAB]  Check the comparison result: %s" % conclusion)  # logger - check the comparison result: SUCCESS/FAILED
        print("[GRAB]  Check the comparison result: %s" % conclusion)  # print - check the comparison result: SUCCESS/FAILED
        self.report_to_xlsx(testcase=testcase,
                            col=self.app.testintex,
                            report=self.app.report_file,
                            result=conclusion)
        self.app.write_log("[EXCEL] Fill the xlsx report")  # logger - fill the xlsx report
        print("[EXCEL] Fill the xlsx report")  # print - fill the xlsx report
        self.compare_result = result
        return result


class Capture:
    def __init__(self, app):
        self.app = app
        self._thread = KThread(target=self._loop)
        self._handlers = []
        self.cap = cv2.VideoCapture(0)
        self.cap.set(3, 720)
        self.cap.set(4, 576)

    def start(self):
        self._thread.start()
        self.app.write_log("[GRAB]  Launch Grabber. Start capturing")  # logger - Launch Grabber. Start capturing
        print("[GRAB]  Launch Grabber. Start capturing")  # print - Launch Grabber. Start capturing

    def stop(self):
        self._thread.kill()
        self.app.write_log("[GRAB]  Turn off Grabber. Stop capturing")  # logger - Turn off Grabber. Stop capturing
        print("[GRAB]  Turn off Grabber. Stop capturing")  # print - Turn off Grabber. Stop capturing

    def _loop(self):
        while 1:
            ret, frame = self.cap.read()
            for handler in self._handlers:
                handler(ret, frame)

    def register_handler(self, handler):
        self._handlers.append(handler)

    def close_session(self):
        self._handlers = []
        self.cap.release()


